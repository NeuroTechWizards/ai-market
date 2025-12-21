"""Загрузка RFSD данных (Parquet) по годам из Hugging Face."""

from __future__ import annotations

from typing import Iterable, Sequence, Any
import logging

import polars as pl

logger = logging.getLogger(__name__)

_AVAILABLE_YEARS = list(range(2011, 2025))

# Глобальный кэш данных по годам
_data_cache: dict[int, pl.DataFrame] = {}


def list_available_years() -> list[int]:
    """Возвращает список доступных годов."""
    return _AVAILABLE_YEARS.copy()


def _validate_year(year: int) -> None:
    if year not in _AVAILABLE_YEARS:
        raise ValueError(f"Год {year} недоступен. Допустимо: {min(_AVAILABLE_YEARS)}–{max(_AVAILABLE_YEARS)}")


from .settings import settings


def _scan_year(year: int, columns: Sequence[str] | None = None) -> pl.LazyFrame:
    """Возвращает ленивый скан по году с добавленной колонкой year.
    
    Использует кэш если данные для года загружены в память.
    """
    _validate_year(year)
    
    # Проверяем кэш
    if year in _data_cache:
        df = _data_cache[year]
        # Если запрошены конкретные колонки, фильтруем
        if columns is not None:
            available_cols = [col for col in columns if col in df.columns]
            if "year" not in available_cols and "year" in df.columns:
                available_cols.append("year")
            df = df.select(available_cols)
        # Возвращаем как LazyFrame из кэша
        return df.lazy()
    
    # Если нет в кэше, читаем с HF
    path = f"hf://datasets/irlspbru/RFSD/RFSD/year={year}/*.parquet"
    
    storage_options = None
    if settings.HF_TOKEN:
        storage_options = {"token": settings.HF_TOKEN}
        
    scan = pl.scan_parquet(path, storage_options=storage_options)
    scan = scan.with_columns(pl.lit(year).alias("year"))
    if columns is not None:
        scan = scan.select(list(columns))
    return scan


def preload_cache(years: list[int] | None = None) -> None:
    """Предзагружает данные указанных годов в кэш.
    
    Args:
        years: Список годов для кэширования. Если None, использует CACHE_YEARS из настроек.
    """
    if years is None:
        # Парсим CACHE_YEARS из настроек
        cache_years_str = settings.CACHE_YEARS
        try:
            years = [int(y.strip()) for y in cache_years_str.split(",")]
        except ValueError:
            logger.warning(f"Не удалось распарсить CACHE_YEARS: {cache_years_str}")
            return
    
    logger.info(f"🔄 Начинаю предзагрузку кэша для годов: {years}")
    
    for year in years:
        if year not in _AVAILABLE_YEARS:
            logger.warning(f"⚠️ Год {year} недоступен, пропускаю")
            continue
        
        if year in _data_cache:
            logger.info(f"✅ Год {year} уже в кэше, пропускаю")
            continue
        
        try:
            logger.info(f"📥 Загружаю год {year} в кэш...")
            df = load_year(year)
            _data_cache[year] = df
            logger.info(f"✅ Год {year} загружен в кэш ({len(df):,} строк, {df.estimated_size('mb'):.1f} MB)")
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки года {year}: {e}")
    
    total_size_mb = sum(df.estimated_size('mb') for df in _data_cache.values())
    logger.info(f"🎉 Кэш готов! Всего годов: {len(_data_cache)}, размер: {total_size_mb:.1f} MB")


def clear_cache() -> None:
    """Очищает кэш данных."""
    _data_cache.clear()
    logger.info("🗑️ Кэш очищен")


def get_cache_info() -> dict[str, Any]:
    """Возвращает информацию о кэше."""
    return {
        "cached_years": sorted(_data_cache.keys()),
        "total_years": len(_data_cache),
        "total_size_mb": round(sum(df.estimated_size('mb') for df in _data_cache.values()), 2),
        "total_rows": sum(len(df) for df in _data_cache.values())
    }


def load_year(year: int, columns: Sequence[str] | None = None) -> pl.DataFrame:
    """Загружает один год RFSD из Hugging Face Parquet."""
    return _scan_year(year, columns=columns).collect()


def filter_inn_year(
    year: int,
    inn: str,
    columns: Sequence[str],
    limit: int = 200,
) -> pl.DataFrame:
    """Фильтр по ИНН для указанного года."""
    cols = list(columns)
    if "inn" not in cols:
        cols.append("inn")
    return (
        _scan_year(year, columns=cols)
        .filter(pl.col("inn") == inn)
        .limit(limit)
        .collect()
    )


def get_schema_columns(year: int) -> list[str]:
    """Возвращает список доступных колонок для указанного года."""
    _validate_year(year)
    path = f"hf://datasets/irlspbru/RFSD/RFSD/year={year}/*.parquet"
    
    storage_options = None
    if settings.HF_TOKEN:
        storage_options = {"token": settings.HF_TOKEN}
        
    scan = pl.scan_parquet(path, storage_options=storage_options)
    schema = scan.collect_schema()
    columns = list(schema.keys())
    if "year" not in columns:
        columns.append("year")
    return columns


def sample_year(year: int, columns: Sequence[str] | None = None, n: int = 5) -> pl.DataFrame:
    """Возвращает первые n строк указанного года без полного collect."""
    return _scan_year(year, columns=columns).limit(n).collect()


def load_indicators_dict() -> dict[str, str]:
    """Загружает справочник индикаторов из Excel."""
    import os
    import openpyxl
    
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(os.path.dirname(current_dir))
    
    possible_paths = [
        os.path.join(project_root, "docs", "databook", "rfsd_databook_ru.xlsx"),
        os.path.join(project_root, "..", "docs", "databook", "rfsd_databook_ru.xlsx"),
        r"D:\OneDrive\Работа\AI consalting\ИИ-агент инвест дир\Программа и Код\AI market\docs\databook\rfsd_databook_ru.xlsx"
    ]
    
    databook_path = None
    for p in possible_paths:
        if os.path.exists(p):
            databook_path = p
            break
            
    if not databook_path:
        return {}

    try:
        wb = openpyxl.load_workbook(databook_path, read_only=True, data_only=True)
        if 'databook' in wb.sheetnames:
            ws = wb['databook']
        else:
            ws = wb.active
            
        headers = [cell.value for cell in ws[1]]
        try:
            code_idx = headers.index('code')
            name_idx = headers.index('name_ru')
        except ValueError:
            return {}
            
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[code_idx]
            name = row[name_idx]
            if code:
                result[str(code)] = str(name) if name else ""
                
        return result
    except Exception as e:
        print(f"Error loading databook: {e}")
        return {}
