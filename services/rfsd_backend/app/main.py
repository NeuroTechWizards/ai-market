"""Основной модуль FastAPI приложения RFSD Backend."""

import logging
import asyncio
from time import perf_counter
from typing import Any
from datetime import datetime
from io import BytesIO

from fastapi import FastAPI, HTTPException, Query, Response
import polars as pl

from . import schemas
from .rfsd_loader import filter_inn_year, get_schema_columns, sample_year, load_indicators_dict, _scan_year
from .settings import settings

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

app = FastAPI(
    title="RFSD Backend",
    description="Backend сервис для Russian Financial Statements Database",
    version="0.1.0",
)

_DEFAULT_FIELDS = [
    "inn",
    "year",
    "region",
    "okved_section",
    "okved",
    "line_2110",
    "line_2300",
    "line_2400",
]

_INDICATORS_DICT = {}

@app.on_event("startup")
async def startup_event():
    """Загружаем справочники при старте."""
    global _INDICATORS_DICT
    
    # Проверка токена
    if settings.HF_TOKEN:
        masked_token = settings.HF_TOKEN[:4] + "..." + settings.HF_TOKEN[-4:]
        logger.info(f"✅ HF_TOKEN found: {masked_token}")
    else:
        logger.warning("❌ HF_TOKEN NOT found! Requests might be rate-limited.")
        
    logger.info("Loading indicators databook...")
    _INDICATORS_DICT = load_indicators_dict()
    logger.info(f"Loaded {len(_INDICATORS_DICT)} indicators.")

from .agent_orchestrator import agent

@app.post("/agent/query", response_model=schemas.AgentQueryResponse)
async def agent_query(request: schemas.AgentQueryRequest) -> schemas.AgentQueryResponse:
    """Эндпоинт для AI-агента (аналитика текста)."""
    start_time = perf_counter()
    logger.info(f"Agent query: {request.query}")
    
    answer = await agent.process_query(request.query)
    
    elapsed_ms = (perf_counter() - start_time) * 1000
    
    return schemas.AgentQueryResponse(
        answer=answer,
        meta={
            "elapsed_ms": round(elapsed_ms, 2)
        }
    )
async def health_check() -> dict[str, str]:
    """Проверка здоровья сервиса."""
    return {"status": "ok"}


@app.get("/rfsd/sample")
async def get_sample(
    year: int = Query(default=2023, ge=2011, le=2024, description="Год данных"),
    limit: int = Query(default=5, ge=1, le=100, description="Количество строк (1-100)"),
    fields: str | None = Query(default="inn,year", description="Список полей через запятую"),
) -> dict[str, Any]:
    """Быстрый endpoint для получения сэмпла данных."""
    if fields:
        fields_list = [f.strip() for f in fields.split(",") if f.strip()]
    else:
        fields_list = ["inn", "year"]

    try:
        df = sample_year(year=year, columns=fields_list, n=limit)
        rows = df.to_dicts()
        return {
            "year": year,
            "columns": fields_list,
            "rows": rows,
        }
    except Exception as e:
        logger.error(f"Ошибка при получении sample: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/rfsd/company_timeseries", response_model=schemas.TableResponse)
async def company_timeseries(request: schemas.CompanyTimeseriesRequest) -> schemas.TableResponse:
    """Поиск компании по ИНН в нескольких годах."""
    years_to_scan = request.years if request.years is not None else [2022, 2023, 2024]
    fields = request.fields if request.fields is not None else _DEFAULT_FIELDS.copy()

    if len(years_to_scan) > 1 and "year" not in fields:
        fields.append("year")

    dropped_fields: list[str] = []
    if years_to_scan:
        schema_columns = get_schema_columns(years_to_scan[0])
        valid_fields = [f for f in fields if f in schema_columns or f == "year"]
        dropped_fields = [f for f in fields if f not in valid_fields]
        fields = valid_fields

    logger.info(f"company_timeseries: inn={request.inn}, years={years_to_scan}, fields={fields}, limit={request.limit}")

    start_time = perf_counter()
    frames: list = []
    per_year_elapsed_ms: dict[int, float] = {}

    for year in years_to_scan:
        year_start = perf_counter()
        try:
            df = filter_inn_year(year=year, inn=request.inn, columns=fields, limit=request.limit)
            if df.height > 0:
                frames.append(df)
            per_year_elapsed_ms[year] = (perf_counter() - year_start) * 1000
        except Exception as e:
            logger.warning(f"Ошибка при обработке года {year}: {e}")
            per_year_elapsed_ms[year] = (perf_counter() - year_start) * 1000

    if frames:
        if len(frames) > 1:
            result_df = pl.concat(frames, how="diagonal")
        else:
            result_df = frames[0]

        if "year" in result_df.columns:
            result_df = result_df.sort("year")

        if result_df.height > request.limit:
            result_df = result_df.head(request.limit)

        columns = result_df.columns
        rows = result_df.to_dicts()
        matched_rows = result_df.height
    else:
        columns = fields
        rows = []
        matched_rows = 0

    elapsed_ms = (perf_counter() - start_time) * 1000

    meta = {
        "years_scanned": years_to_scan,
        "matched_rows": matched_rows,
        "elapsed_ms": round(elapsed_ms, 2),
        "per_year_elapsed_ms": {str(k): round(v, 2) for k, v in per_year_elapsed_ms.items()},
    }

    if dropped_fields:
        meta["dropped_fields"] = dropped_fields

    return schemas.TableResponse(columns=columns, rows=rows, meta=meta, files=None)


@app.post("/rfsd/company_revenue_timeseries", response_model=schemas.CompanyRevenueTimeseriesResponse)
async def company_revenue_timeseries(request: schemas.CompanyRevenueTimeseriesRequest) -> schemas.CompanyRevenueTimeseriesResponse:
    """Вернуть выручку компании (line_2110) по годам."""
    years_to_scan = request.years if request.years is not None else [2019, 2020, 2021, 2022, 2023]
    logger.info(f"company_revenue_timeseries: inn={request.inn}, years={years_to_scan}")
    
    start_time = perf_counter()
    series: list[schemas.RevenueYear] = []
    matched_rows = 0
    per_year_elapsed_ms: dict[int, float] = {}
    
    fields = ["inn", "year", "line_2110"]
    
    for year in years_to_scan:
        year_start = perf_counter()
        try:
            df = filter_inn_year(year=year, inn=request.inn, columns=fields, limit=1)
            
            if df.height > 0:
                row = df.row(0, named=True)
                revenue_val = row.get("line_2110")
                if revenue_val is not None:
                    try:
                        revenue_val = float(revenue_val)
                    except (ValueError, TypeError):
                        revenue_val = None
                        
                series.append(schemas.RevenueYear(year=year, revenue=revenue_val))
                matched_rows += 1
                
            per_year_elapsed_ms[year] = (perf_counter() - year_start) * 1000
        except Exception as e:
            logger.warning(f"Ошибка при получении выручки за {year}: {e}")
            per_year_elapsed_ms[year] = (perf_counter() - year_start) * 1000

    series.sort(key=lambda x: x.year)
    elapsed_ms = (perf_counter() - start_time) * 1000
    
    return schemas.CompanyRevenueTimeseriesResponse(
        inn=request.inn,
        series=series,
        meta={
            "years_scanned": years_to_scan,
            "matched_rows": matched_rows,
            "elapsed_ms": round(elapsed_ms, 2),
            "per_year_elapsed_ms": {str(k): round(v, 2) for k, v in per_year_elapsed_ms.items()},
        }
    )


@app.post("/rfsd/export_company_revenue_xlsx")
async def export_company_revenue_xlsx(request: schemas.ExportCompanyRevenueRequest):
    """Экспорт выручки компании в Excel файл."""
    internal_req = schemas.CompanyRevenueTimeseriesRequest(inn=request.inn, years=request.years)
    data = await company_revenue_timeseries(internal_req)
    
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "revenue_timeseries"
        ws1.append(["year", "revenue"])
        for item in data.series:
            ws1.append([item.year, item.revenue])
            
        ws2 = wb.create_sheet("meta")
        ws2.append(["inn", "elapsed_ms", "generated_at"])
        ws2.append([data.inn, data.meta.get("elapsed_ms"), datetime.now().isoformat()])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"rfsd_revenue_{request.inn}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Ошибка при генерации Excel: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Ошибка генерации Excel: {str(e)}")


@app.post("/rfsd/export_full_profile_xlsx")
async def export_full_profile_xlsx(request: schemas.ExportFullProfileRequest):
    """Экспорт ВСЕХ финансовых показателей компании в Excel (Profile)."""
    years_to_scan = request.years if request.years is not None else [2019, 2020, 2021, 2022, 2023]
    logger.info(f"export_full_profile_xlsx: inn={request.inn}, years={years_to_scan}")
    
    start_time = perf_counter()

    try:
        schema_cols = get_schema_columns(2023)
        financial_lines = [c for c in schema_cols if c.startswith("line_")]
        financial_lines.sort() 
        
        fields_to_load = ["inn", "year"] + financial_lines
        
        frames = []
        for year in years_to_scan:
            try:
                current_year_schema = get_schema_columns(year)
                current_fields = [f for f in fields_to_load if f in current_year_schema or f == "year"]
                
                df = filter_inn_year(year=year, inn=request.inn, columns=current_fields, limit=1)
                if df.height > 0:
                    frames.append(df)
            except Exception as e:
                logger.warning(f"Error loading year {year}: {e}")
        
        if not frames:
             raise HTTPException(status_code=404, detail="Data not found for this INN")
             
        full_df = pl.concat(frames, how="diagonal")
        full_df = full_df.sort("year")
        
        matrix = {code: {} for code in financial_lines}
        
        for row in full_df.to_dicts():
            y = row["year"]
            for k, v in row.items():
                if k.startswith("line_") and v is not None:
                     matrix[k][y] = v
                     
        indicators_dict = _INDICATORS_DICT
                     
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Profile"
        
        headers = ["Indicator Code", "Indicator Name (RU)"] + [str(y) for y in years_to_scan]
        ws.append(headers)
        
        for code in financial_lines:
            name_ru = indicators_dict.get(code, "")
            row_data = [code, name_ru]
            for year in years_to_scan:
                val = matrix.get(code, {}).get(year)
                row_data.append(val)
            ws.append(row_data)
            
        ws2 = wb.create_sheet("meta")
        ws2.append(["inn", "generated_at", "elapsed_s"])
        elapsed = perf_counter() - start_time
        ws2.append([request.inn, datetime.now().isoformat(), round(elapsed, 2)])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"rfsd_profile_{request.inn}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Full profile export error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/rfsd/sector_benchmark", response_model=schemas.TableResponse)
async def sector_benchmark(request: schemas.SectorBenchmarkRequest) -> schemas.TableResponse:
    """Сравнение компании с отраслью (медиана) - ОПТИМИЗИРОВАННАЯ ВЕРСИЯ."""
    
    years_to_scan = request.years if request.years is not None else [2019, 2020, 2021, 2022, 2023]
    metrics = request.metrics if request.metrics is not None else ["line_2110", "line_2400"]
    
    logger.info(f"sector_benchmark: inn={request.inn}, years={years_to_scan}, metrics={metrics}")
    
    start_time = perf_counter()
    per_year_elapsed_ms: dict[str, float] = {}
    result_rows = []
    rate_limit_errors = []
    
    for idx, year in enumerate(years_to_scan):
        # Добавляем задержку между запросами к Hugging Face для снижения rate limiting
        if idx > 0:
            await asyncio.sleep(2.0)  # 2 секунды задержка между годами
        
        year_start = perf_counter()
        try:
            # Шаг A: Находим компанию и её отрасль
            comp_df = filter_inn_year(
                year=year,
                inn=request.inn,
                columns=["inn", "year", "okved_section"] + metrics,
                limit=1
            )
            
            if comp_df.height == 0:
                per_year_elapsed_ms[str(year)] = (perf_counter() - year_start) * 1000
                continue
                
            comp_row = comp_df.row(0, named=True)
            section = comp_row.get("okved_section")
            
            if not section:
                logger.warning(f"No okved_section for INN {request.inn} in {year}")
                per_year_elapsed_ms[str(year)] = (perf_counter() - year_start) * 1000
                continue
                
            # Шаг B: Оптимизированный расчёт статистики сектора
            # УСКОРИТЕЛЬ 1: Только нужные колонки
            sector_cols = ["okved_section"] + metrics
            
            # УСКОРИТЕЛЬ 2: Lazy evaluation с агрегацией
            scan = _scan_year(year, columns=sector_cols)
            
            # Фильтруем по секции
            scan = scan.filter(pl.col("okved_section") == section)
            
            # УСКОРИТЕЛЬ 3: Ограничение выборки (предохранитель)
            scan = scan.head(request.limit_sector)
            
            # Фильтруем null по каждой метрике
            for m in metrics:
                scan = scan.filter(pl.col(m).is_not_null())
            
            # УСКОРИТЕЛЬ 4: Только median (без P25/P75 для скорости)
            # Используем quantile с interpolation="nearest" для скорости
            agg_exprs = [pl.len().alias("sector_count")]
            for m in metrics:
                # median через quantile(0.5)
                agg_exprs.append(
                    pl.col(m).quantile(0.5, interpolation="nearest").alias(f"sector_median_{m}")
                )
            
            sector_stats = scan.select(agg_exprs).collect()
            
            if sector_stats.height > 0:
                stats_row = sector_stats.row(0, named=True)
                
                # Формируем итоговую строку
                row_out = {
                    "year": year,
                    "okved_section": section,
                    "sector_count": stats_row["sector_count"],
                    "sampled_rows": min(request.limit_sector, stats_row["sector_count"])
                }
                
                # Данные компании и сектора
                for m in metrics:
                    row_out[f"company_{m}"] = comp_row.get(m)
                    row_out[f"sector_median_{m}"] = stats_row.get(f"sector_median_{m}")
                    
                result_rows.append(row_out)
                
            per_year_elapsed_ms[str(year)] = (perf_counter() - year_start) * 1000
            
        except Exception as e:
            error_msg = str(e)
            # Проверяем, является ли это ошибкой rate limiting
            if "429" in error_msg or "rate limit" in error_msg.lower():
                rate_limit_errors.append(year)
                logger.warning(f"Rate limit error for year {year}. Skipping.")
            else:
                logger.error(f"Error benchmarking {year}: {e}", exc_info=True)
            per_year_elapsed_ms[str(year)] = (perf_counter() - year_start) * 1000

    elapsed_ms = (perf_counter() - start_time) * 1000
    
    # Сортируем result_rows по году перед возвратом
    result_rows.sort(key=lambda x: x["year"])
    
    # Формируем колонки
    columns = ["year", "okved_section", "sector_count", "sampled_rows"]
    for m in metrics:
        columns.extend([f"company_{m}", f"sector_median_{m}"])

    meta = {
        "years_scanned": years_to_scan,
        "matched_rows": len(result_rows),
        "elapsed_ms": round(elapsed_ms, 2),
        "per_year_elapsed_ms": {k: round(v, 2) for k, v in per_year_elapsed_ms.items()}
    }
    
    # Добавляем информацию о rate limiting в мета
    if rate_limit_errors:
        meta["rate_limit_errors"] = rate_limit_errors
        meta["warning"] = f"Некоторые годы ({rate_limit_errors}) не обработаны из-за rate limiting. Попробуйте запросить меньше лет или подождите."

    return schemas.TableResponse(
        columns=columns,
        rows=result_rows,
        meta=meta
    )
